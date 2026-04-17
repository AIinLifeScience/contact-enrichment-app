"""
Enrichment Results Database
============================
SQLite-Datenbank zum automatischen Speichern aller Enrichment-Ergebnisse.
- Auto-Save nach jedem Kontakt (kein Datenverlust mehr)
- Cache-Lookup: Bereits recherchierte Kontakte werden sofort angezeigt
- Merge-Logik: Neue Infos werden ergänzt, nie überschrieben
- Delta-Scan-ready: raw_findings + monitoring_tags für zukünftige Scans
- Export: Alle Ergebnisse als formatierte Excel-Datei
"""

import json
import os
import re
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional


# Alle Result-Felder die aus enrich_contact() kommen
RESULT_FIELDS = [
    "email", "email_status", "email_verifizierung",
    "referenz_email", "referenz_email_quelle",
    "personal_phone", "company_phone",
    "conferences", "podcasts_videos", "job_changes",
    "linkedin_activity", "birthday",
    "relevant_info", "zusammenfassung",
    "personalisierte_nachricht", "kanal_empfehlung",
    "monitoring_tags",
]

# Felder die als JSON gespeichert werden (Listen/Dicts)
JSON_FIELDS = {"sources", "leadgen_match"}

# Felder die bei Re-Run gemergt (ergänzt) statt überschrieben werden
# → Textfelder wo neue Infos zu alten hinzukommen sollen
MERGE_FIELDS = {
    "conferences", "podcasts_videos", "job_changes",
    "linkedin_activity", "relevant_info", "monitoring_tags",
    "raw_findings",
}

# Felder die bei Re-Run IMMER überschrieben werden (neueste Version gewinnt)
# → Email, Telefon, Zusammenfassung, Nachricht = immer aktuellste Version
OVERWRITE_FIELDS = {
    "email", "email_status", "email_verifizierung",
    "referenz_email", "referenz_email_quelle",
    "personal_phone", "company_phone", "birthday",
    "zusammenfassung", "personalisierte_nachricht", "kanal_empfehlung",
}


def _merge_text_field(old_value: str, new_value: str) -> str:
    """Mergt zwei Textfelder: Ergänzt neue Infos, vermeidet Duplikate.

    Strategie:
    - Splittet beide Werte in einzelne Info-Einträge (getrennt durch Newline oder ' | ')
    - Prüft für jeden neuen Eintrag ob er schon im alten vorkommt (fuzzy, 80% Overlap)
    - Hängt nur wirklich neue Einträge an, mit Datum-Tag
    """
    if not new_value or new_value.strip() in ("", "Nicht gefunden", "-", "N/A"):
        return old_value or ""
    if not old_value or old_value.strip() in ("", "Nicht gefunden", "-", "N/A"):
        return new_value

    # Einträge extrahieren (getrennt durch Newline, ' | ', oder '; ')
    def _split_entries(text):
        entries = re.split(r'\n|(?:\s*\|\s*)| ;\s*', text.strip())
        return [e.strip() for e in entries if e.strip()]

    old_entries = _split_entries(old_value)
    new_entries = _split_entries(new_value)

    # Normalisierte Versionen für Duplikat-Check
    def _normalize(text):
        return re.sub(r'\s+', ' ', text.lower().strip())

    old_normalized = {_normalize(e) for e in old_entries}

    # Nur wirklich neue Einträge hinzufügen
    added = []
    for entry in new_entries:
        norm = _normalize(entry)
        # Fuzzy-Check: Ist >80% des Texts schon in einem alten Eintrag?
        is_duplicate = False
        for old_norm in old_normalized:
            if norm == old_norm:
                is_duplicate = True
                break
            # Substring-Check: Neuer Eintrag ist Teil eines alten (oder umgekehrt)
            if len(norm) > 10 and (norm in old_norm or old_norm in norm):
                is_duplicate = True
                break
        if not is_duplicate:
            added.append(entry)

    if not added:
        return old_value  # Nichts Neues gefunden

    # Datum-Tag für neue Einträge
    date_tag = datetime.now().strftime("[Neu %d.%m.%Y]")

    # Zusammenführen: Alte Einträge + neue mit Datum-Tag
    merged = old_value.rstrip()
    merged += f"\n{date_tag} " + " | ".join(added)
    return merged


class EnrichmentDB:
    """SQLite-Datenbank für Enrichment-Ergebnisse."""

    DEFAULT_DB_PATH = str(
        Path(__file__).parent / "data" / "enrichment_results.db"
    )

    def __init__(self, db_path: Optional[str] = None):
        self.db_path = (
            db_path
            or os.environ.get("ENRICHMENT_DB_PATH")
            or self.DEFAULT_DB_PATH
        )
        # Verzeichnis anlegen falls nötig
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        self._init_db()

    def _get_conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path, timeout=5.0)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")  # Concurrent reads
        return conn

    def _init_db(self):
        """Erstellt die Tabelle falls sie nicht existiert."""
        conn = self._get_conn()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS enrichment_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,

                -- Lookup-Keys
                name TEXT NOT NULL,
                company TEXT NOT NULL DEFAULT '',

                -- Input-Kontext
                title TEXT DEFAULT '',
                location TEXT DEFAULT '',

                -- Enrichment-Ergebnisse
                email TEXT DEFAULT '',
                email_status TEXT DEFAULT '',
                email_verifizierung TEXT DEFAULT '',
                referenz_email TEXT DEFAULT '',
                referenz_email_quelle TEXT DEFAULT '',
                personal_phone TEXT DEFAULT '',
                company_phone TEXT DEFAULT '',
                conferences TEXT DEFAULT '',
                podcasts_videos TEXT DEFAULT '',
                job_changes TEXT DEFAULT '',
                linkedin_activity TEXT DEFAULT '',
                birthday TEXT DEFAULT '',
                relevant_info TEXT DEFAULT '',
                zusammenfassung TEXT DEFAULT '',
                personalisierte_nachricht TEXT DEFAULT '',
                kanal_empfehlung TEXT DEFAULT '',
                monitoring_tags TEXT DEFAULT '',

                -- Delta-Scan-Daten
                raw_findings TEXT DEFAULT '',
                sources TEXT DEFAULT '[]',
                leadgen_match TEXT DEFAULT '{}',

                -- Metadata
                llm_provider TEXT DEFAULT '',
                search_depth INTEGER DEFAULT 6,
                duration_seconds REAL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        # Unique Index für Cache-Lookup (case-insensitive)
        conn.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_name_company
            ON enrichment_results (name COLLATE NOCASE, company COLLATE NOCASE)
        """)
        conn.commit()
        conn.close()

    def lookup(self, name: str, company: str = "") -> Optional[dict]:
        """Cache-Lookup: Gibt gespeichertes Ergebnis zurück oder None.

        Match: name + company (case-insensitive).
        """
        if not name:
            return None
        conn = self._get_conn()
        row = conn.execute(
            """SELECT * FROM enrichment_results
               WHERE name = ? COLLATE NOCASE
                 AND company = ? COLLATE NOCASE""",
            (name.strip(), (company or "").strip()),
        ).fetchone()
        conn.close()

        if not row:
            return None

        result = dict(row)
        # JSON-Felder deserialisieren
        for field in JSON_FIELDS:
            if field in result and isinstance(result[field], str):
                try:
                    result[field] = json.loads(result[field])
                except (json.JSONDecodeError, TypeError):
                    pass
        # Alter berechnen
        try:
            created = datetime.fromisoformat(result["created_at"])
            result["_age_days"] = (datetime.now() - created).days
            result["_age_label"] = self._age_label(result["_age_days"])
        except (ValueError, TypeError):
            result["_age_days"] = 0
            result["_age_label"] = "heute"

        return result

    def save(self, name: str, company: str, result: dict,
             metadata: Optional[dict] = None) -> int:
        """Speichert ein Enrichment-Ergebnis — MERGE statt Überschreiben.

        Bei bestehendem Eintrag (gleicher Name+Firma):
        - MERGE_FIELDS (Konferenzen, Podcasts, LinkedIn etc.) werden ERGÄNZT
        - OVERWRITE_FIELDS (Email, Zusammenfassung, Nachricht) werden aktualisiert
        - created_at bleibt erhalten, updated_at wird aktualisiert
        - Neue Infos bekommen ein Datum-Tag [Neu DD.MM.YYYY]
        """
        now = datetime.now().isoformat()
        meta = metadata or {}

        # JSON-Felder serialisieren
        sources = result.get("sources", [])
        if isinstance(sources, (list, dict)):
            sources = json.dumps(sources, ensure_ascii=False)
        leadgen = result.get("leadgen_match", {})
        if isinstance(leadgen, dict):
            leadgen = json.dumps(leadgen, ensure_ascii=False)

        conn = self._get_conn()

        # Prüfen ob schon ein Eintrag existiert
        existing = conn.execute(
            """SELECT * FROM enrichment_results
               WHERE name = ? COLLATE NOCASE
                 AND company = ? COLLATE NOCASE""",
            (name.strip(), (company or "").strip()),
        ).fetchone()

        if existing:
            # --- MERGE-Modus: Bestehenden Eintrag ergänzen ---
            old = dict(existing)
            created_at = old["created_at"]

            # Merge-Felder: Neue Infos an bestehende anhängen
            merged = {}
            for field in MERGE_FIELDS:
                old_val = old.get(field, "") or ""
                new_val = result.get(field, "") or ""
                merged[field] = _merge_text_field(old_val, new_val)

            # Sources mergen (JSON-Liste)
            old_sources = old.get("sources", "[]")
            if isinstance(old_sources, str):
                try:
                    old_sources = json.loads(old_sources)
                except (json.JSONDecodeError, TypeError):
                    old_sources = []
            new_sources = result.get("sources", [])
            if isinstance(new_sources, str):
                try:
                    new_sources = json.loads(new_sources)
                except (json.JSONDecodeError, TypeError):
                    new_sources = []
            # Nur neue Sources hinzufügen
            existing_urls = {s.get("url", s) if isinstance(s, dict) else s for s in old_sources}
            for s in new_sources:
                url = s.get("url", s) if isinstance(s, dict) else s
                if url not in existing_urls:
                    old_sources.append(s)
            sources = json.dumps(old_sources, ensure_ascii=False)

            conn.execute("""
                UPDATE enrichment_results SET
                    title = ?, location = ?,
                    email = ?, email_status = ?, email_verifizierung = ?,
                    referenz_email = ?, referenz_email_quelle = ?,
                    personal_phone = ?, company_phone = ?,
                    conferences = ?, podcasts_videos = ?, job_changes = ?,
                    linkedin_activity = ?, birthday = ?,
                    relevant_info = ?, zusammenfassung = ?,
                    personalisierte_nachricht = ?, kanal_empfehlung = ?,
                    monitoring_tags = ?, raw_findings = ?,
                    sources = ?, leadgen_match = ?,
                    llm_provider = ?, search_depth = ?, duration_seconds = ?,
                    updated_at = ?
                WHERE id = ?
            """, (
                result.get("_title", meta.get("title", old.get("title", ""))),
                result.get("_location", meta.get("location", old.get("location", ""))),
                # Overwrite-Felder: Neuer Wert gewinnt (wenn nicht leer)
                result.get("email", "") or old.get("email", ""),
                result.get("email_status", "") or old.get("email_status", ""),
                result.get("email_verifizierung", "") or old.get("email_verifizierung", ""),
                result.get("referenz_email", "") or old.get("referenz_email", ""),
                result.get("referenz_email_quelle", "") or old.get("referenz_email_quelle", ""),
                result.get("personal_phone", "") or old.get("personal_phone", ""),
                result.get("company_phone", "") or old.get("company_phone", ""),
                # Merge-Felder: Ergänzt statt überschrieben
                merged.get("conferences", ""),
                merged.get("podcasts_videos", ""),
                merged.get("job_changes", ""),
                merged.get("linkedin_activity", ""),
                result.get("birthday", "") or old.get("birthday", ""),
                merged.get("relevant_info", ""),
                # Zusammenfassung + Nachricht: Immer neueste Version
                result.get("zusammenfassung", "") or old.get("zusammenfassung", ""),
                result.get("personalisierte_nachricht", "") or old.get("personalisierte_nachricht", ""),
                result.get("kanal_empfehlung", "") or old.get("kanal_empfehlung", ""),
                merged.get("monitoring_tags", ""),
                merged.get("raw_findings", ""),
                sources,
                leadgen,
                meta.get("llm_provider", old.get("llm_provider", "")),
                meta.get("search_depth", old.get("search_depth", 6)),
                meta.get("duration_seconds", 0),
                now,
                old["id"],
            ))
            rowid = old["id"]
            n_merged = sum(1 for f in MERGE_FIELDS
                           if merged.get(f, "") != (old.get(f, "") or ""))
            if n_merged:
                print(f"  [DB MERGE] {name}: {n_merged} Felder ergänzt (nicht überschrieben)")
        else:
            # --- INSERT-Modus: Neuer Kontakt ---
            created_at = now
            conn.execute("""
                INSERT INTO enrichment_results (
                    name, company, title, location,
                    email, email_status, email_verifizierung,
                    referenz_email, referenz_email_quelle,
                    personal_phone, company_phone,
                    conferences, podcasts_videos, job_changes,
                    linkedin_activity, birthday,
                    relevant_info, zusammenfassung,
                    personalisierte_nachricht, kanal_empfehlung,
                    monitoring_tags, raw_findings, sources, leadgen_match,
                    llm_provider, search_depth, duration_seconds,
                    created_at, updated_at
                ) VALUES (
                    ?, ?, ?, ?,
                    ?, ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?, ?, ?,
                    ?, ?, ?,
                    ?, ?
                )
            """, (
                name.strip(), (company or "").strip(),
                result.get("_title", meta.get("title", "")),
                result.get("_location", meta.get("location", "")),
                result.get("email", ""),
                result.get("email_status", ""),
                result.get("email_verifizierung", ""),
                result.get("referenz_email", ""),
                result.get("referenz_email_quelle", ""),
                result.get("personal_phone", ""),
                result.get("company_phone", ""),
                result.get("conferences", ""),
                result.get("podcasts_videos", ""),
                result.get("job_changes", ""),
                result.get("linkedin_activity", ""),
                result.get("birthday", ""),
                result.get("relevant_info", ""),
                result.get("zusammenfassung", ""),
                result.get("personalisierte_nachricht", ""),
                result.get("kanal_empfehlung", ""),
                result.get("monitoring_tags", ""),
                result.get("raw_findings", ""),
                sources,
                leadgen,
                meta.get("llm_provider", ""),
                meta.get("search_depth", 6),
                meta.get("duration_seconds", 0),
                created_at,
                now,
            ))
            rowid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        conn.commit()
        conn.close()
        return rowid

    def all_results(self) -> list[dict]:
        """Gibt alle gespeicherten Ergebnisse zurück (neueste zuerst)."""
        conn = self._get_conn()
        rows = conn.execute(
            "SELECT * FROM enrichment_results ORDER BY updated_at DESC"
        ).fetchall()
        conn.close()
        results = []
        for row in rows:
            d = dict(row)
            try:
                created = datetime.fromisoformat(d["created_at"])
                d["_age_days"] = (datetime.now() - created).days
                d["_age_label"] = self._age_label(d["_age_days"])
            except (ValueError, TypeError):
                d["_age_days"] = 0
                d["_age_label"] = "?"
            results.append(d)
        return results

    def count(self) -> int:
        """Anzahl gespeicherter Enrichments."""
        conn = self._get_conn()
        n = conn.execute("SELECT COUNT(*) FROM enrichment_results").fetchone()[0]
        conn.close()
        return n

    def export_excel(self, output_path: str):
        """Exportiert alle Ergebnisse als formatierte Excel-Datei."""
        import pandas as pd

        results = self.all_results()
        if not results:
            return None

        # Nur relevante Spalten, in sinnvoller Reihenfolge
        export_cols = [
            "name", "company", "title", "location",
            "email", "email_status", "email_verifizierung",
            "personal_phone", "company_phone",
            "conferences", "podcasts_videos", "job_changes",
            "linkedin_activity", "relevant_info",
            "zusammenfassung", "personalisierte_nachricht",
            "kanal_empfehlung", "monitoring_tags",
            "llm_provider", "search_depth", "duration_seconds",
            "created_at", "updated_at",
        ]
        # Interne Felder entfernen
        clean = []
        for r in results:
            clean.append({k: r.get(k, "") for k in export_cols})

        df = pd.DataFrame(clean)

        # Spalten-Header umbenennen für bessere Lesbarkeit
        rename = {
            "name": "Name", "company": "Firma", "title": "Titel",
            "location": "Standort", "email": "E-Mail",
            "email_status": "E-Mail Status",
            "email_verifizierung": "E-Mail Verifizierung",
            "personal_phone": "Persönliches Telefon",
            "company_phone": "Firmentelefon",
            "conferences": "Konferenzen",
            "podcasts_videos": "Podcasts / Videos",
            "job_changes": "Jobwechsel / Karriere",
            "linkedin_activity": "LinkedIn-Aktivität",
            "relevant_info": "Relevante Infos",
            "zusammenfassung": "KI-Zusammenfassung",
            "personalisierte_nachricht": "Personalisierte Nachricht",
            "kanal_empfehlung": "Kanal-Empfehlung",
            "monitoring_tags": "Monitoring-Tags",
            "llm_provider": "LLM Provider",
            "search_depth": "Suchtiefe",
            "duration_seconds": "Dauer (Sek)",
            "created_at": "Erstellt am",
            "updated_at": "Aktualisiert am",
        }
        df.rename(columns=rename, inplace=True)
        df.to_excel(output_path, index=False, engine="openpyxl")

        # CI-Formatierung (optional, falls openpyxl vorhanden)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.load_workbook(output_path)
            ws = wb.active

            header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
            header_fill = PatternFill(start_color="370C7B", end_color="370C7B", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = thin_border

            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            for col_idx in range(1, len(df.columns) + 1):
                letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[letter].width = 25

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}1"
            wb.save(output_path)
        except Exception:
            pass  # Formatierung ist nice-to-have, kein Blocker

        return output_path

    def delete(self, row_id: int):
        """Löscht einen Eintrag nach ID."""
        conn = self._get_conn()
        conn.execute("DELETE FROM enrichment_results WHERE id = ?", (row_id,))
        conn.commit()
        conn.close()

    @staticmethod
    def _age_label(days: int) -> str:
        if days == 0:
            return "heute"
        elif days == 1:
            return "gestern"
        elif days < 7:
            return f"vor {days} Tagen"
        elif days < 30:
            weeks = days // 7
            return f"vor {weeks} Woche{'n' if weeks > 1 else ''}"
        else:
            months = days // 30
            return f"vor {months} Monat{'en' if months > 1 else ''}"
