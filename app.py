"""
Contact Data Enrichment App
============================
Streamlit-App: Lädt Kontaktliste, geht automatisch ALLE Kontakte durch,
sucht E-Mail, Telefon, Firmentelefon, Konferenzen, Podcasts, Jobwechsel etc.
und speichert alles in eine neue angereicherte Excel-Datei.
"""

import streamlit as st
import pandas as pd
import os
import time
from io import BytesIO
from datetime import datetime
from pathlib import Path

from enrichment_engine import EnrichmentEngine
from enrichment_db import EnrichmentDB

# --- .env Datei laden (lokale API Keys) ---
def _load_env():
    env_path = Path(__file__).parent / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if "=" in line and not line.startswith("#"):
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip())
_load_env()

# --- Page Config ---
st.set_page_config(
    page_title="Contact Enrichment — Selina Gaertner Consulting",
    page_icon="🧬",
    layout="wide",
)

# --- Corporate Identity ---
CI_PURPLE = "#370C7B"
CI_GOLD = "#FCB02F"
CI_CHARCOAL = "#2D2D2D"
CI_LIGHT_GREY = "#F5F5F5"

st.markdown(f"""
<style>
    /* Header */
    .stApp header {{
        background-color: {CI_PURPLE} !important;
    }}
    /* Sidebar */
    section[data-testid="stSidebar"] {{
        background-color: {CI_LIGHT_GREY} !important;
        border-right: 3px solid {CI_GOLD};
    }}
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3 {{
        color: {CI_PURPLE} !important;
    }}
    /* Haupttitel */
    h1 {{
        color: {CI_PURPLE} !important;
        font-family: Arial, sans-serif !important;
    }}
    h2, h3 {{
        color: {CI_PURPLE} !important;
        font-family: Arial, sans-serif !important;
    }}
    /* Primary Buttons */
    .stButton > button[kind="primary"],
    button[data-testid="stBaseButton-primary"] {{
        background-color: {CI_GOLD} !important;
        color: {CI_CHARCOAL} !important;
        border: none !important;
        font-weight: bold !important;
        font-family: Arial, sans-serif !important;
    }}
    .stButton > button[kind="primary"]:hover,
    button[data-testid="stBaseButton-primary"]:hover {{
        background-color: {CI_PURPLE} !important;
        color: white !important;
    }}
    /* Secondary Buttons */
    .stButton > button {{
        border-color: {CI_PURPLE} !important;
        color: {CI_PURPLE} !important;
        font-family: Arial, sans-serif !important;
    }}
    /* Metrics */
    [data-testid="stMetricValue"] {{
        color: {CI_PURPLE} !important;
        font-family: Arial, sans-serif !important;
    }}
    /* Progress bar */
    .stProgress > div > div > div {{
        background-color: {CI_GOLD} !important;
    }}
    /* Radio buttons */
    .stRadio > div > label > div:first-child {{
        color: {CI_PURPLE} !important;
    }}
    /* Info boxes */
    .stAlert {{
        font-family: Arial, sans-serif !important;
    }}
    /* Gold accent line under title */
    .ci-accent {{
        height: 3px;
        background: linear-gradient(90deg, {CI_GOLD}, {CI_PURPLE});
        margin: -10px 0 20px 0;
        border-radius: 2px;
    }}
    /* Footer */
    .ci-footer {{
        text-align: center;
        color: {CI_CHARCOAL}99;
        font-size: 12px;
        font-family: Arial, sans-serif;
        padding: 20px 0;
        border-top: 2px solid {CI_GOLD};
        margin-top: 40px;
    }}
    /* Body font — span ausschließen, damit Material-Icons (arrow_down etc.) erhalten bleiben */
    .stMarkdown, .stText, p, li {{
        font-family: Arial, sans-serif !important;
    }}
    /* Material Icons NICHT überschreiben — sonst erscheint Icon-Name als Text */
    span[class*="material-icons"],
    span[class*="material-symbols"],
    [data-testid="stIconMaterial"],
    [data-testid="stExpanderToggleIcon"] {{
        font-family: "Material Symbols Rounded", "Material Icons" !important;
    }}
</style>
""", unsafe_allow_html=True)

# --- Constants (konfigurierbar via .env) ---
DEFAULT_EXCEL = os.environ.get(
    "DEFAULT_EXCEL_PATH",
    str(Path.home() / "Documents" / "AI Products I build" / "selina-ai-leadgen" / "kontakte.xlsx"),
)
DEFAULT_OUTPUT_DIR = os.environ.get(
    "DEFAULT_OUTPUT_DIR",
    str(Path.home() / "Documents"),
)

# Sicherheits-Limits
MAX_UPLOAD_SIZE_MB = 10
MAX_CONTACTS = 500

# Enrichment-Ergebnis-DB (Auto-Save nach jedem Kontakt)
enrichment_db = EnrichmentDB()

NEW_COLUMNS = [
    "E-Mail",
    "E-Mail Status",
    "E-Mail Verifizierung",
    "Referenz-Email",
    "Referenz-Email Quelle",
    "Persönliches Telefon",
    "Firmentelefon (Impressum)",
    "Konferenzen (2025/2026)",
    "Podcasts / Videos",
    "Jobwechsel / Karriere",
    "LinkedIn-Aktivität",
    "Geburtstag",
    "Relevante Infos für Ansprache",
    "KI-Zusammenfassung",
    "Personalisierte Nachricht",
    "Kanal-Empfehlung (Email/LinkedIn)",
    "Monitoring-Tags (Delta-Scan)",
    "Enrichment Datum",
]

# Column name mappings for flexible Excel formats
NAME_COLS = ["Name", "name", "NAME", "Kontakt", "kontakt"]
FIRMA_COLS = ["Firma", "firma", "Company", "company", "Unternehmen"]
TITEL_COLS = ["Titel/Position", "Titel", "Position", "Title", "titel"]
ORT_COLS = ["Kontakt Standort", "Standort", "Location", "Stadt", "Firma Stadt"]


def _sanitize_error(msg: str, max_len: int = 150) -> str:
    """Entfernt sensible Infos aus Fehlermeldungen (API Keys, Pfade, URLs)."""
    import re
    # API Keys maskieren
    msg = re.sub(r'(sk-[a-zA-Z0-9_-]{10,})', '[API_KEY]', msg)
    msg = re.sub(r'(AIza[a-zA-Z0-9_-]{10,})', '[API_KEY]', msg)
    msg = re.sub(r'([a-f0-9]{32,})', '[TOKEN]', msg)
    # Auth-Header maskieren
    msg = re.sub(r'(Bearer\s+)[^\s]+', r'\1[REDACTED]', msg, flags=re.IGNORECASE)
    msg = re.sub(r'(api_key=)[^&\s]+', r'\1[REDACTED]', msg)
    # Home-Pfade anonymisieren
    msg = re.sub(r'/Users/[^/\s]+', '/Users/<user>', msg)
    msg = re.sub(r'/home/[^/\s]+', '/home/<user>', msg)
    # Länge begrenzen
    if len(msg) > max_len:
        msg = msg[:max_len] + "..."
    return msg


def find_column(df, candidates):
    """Findet die erste passende Spalte aus einer Liste von Kandidaten."""
    for col_name in candidates:
        if col_name in df.columns:
            return col_name
    return None


def get_cell_value(row, col_name):
    """Holt einen Wert aus einer Zeile, gibt '' zurück wenn leer/None."""
    if col_name is None:
        return ""
    val = row.get(col_name)
    if pd.isna(val) or val is None or str(val).strip().lower() in ("none", "nan", ""):
        return ""
    return str(val).strip()


def is_already_enriched(row):
    """Prüft ob ein Kontakt bereits angereichert wurde (hat schon E-Mail-Daten)."""
    for col in ["E-Mail", "E-Mail Status", "Enrichment Datum"]:
        if col in row.index:
            val = row.get(col)
            if pd.notna(val) and str(val).strip() and str(val).strip().lower() != "none":
                return True
    return False


def run_enrichment(df, engine, max_searches, skip_enriched, progress_bar, status_container,
                    selected_indices=None, force_rerun_indices=None):
    """Reichert ausgewählte Kontakte an. Auto-Save in DB nach jedem Kontakt.

    force_rerun_indices: Set von Indizes die trotz Cache neu recherchiert werden.
    """
    name_col = find_column(df, NAME_COLS)
    firma_col = find_column(df, FIRMA_COLS)
    titel_col = find_column(df, TITEL_COLS)
    ort_col = find_column(df, ORT_COLS)

    if not name_col:
        st.error("Keine 'Name'-Spalte gefunden!")
        return {}

    force_rerun = force_rerun_indices or set()
    results = {}
    # Nur ausgewählte Zeilen; wenn nichts übergeben → alle
    rows_to_process = [(i, df.iloc[i]) for i in (selected_indices if selected_indices is not None else range(len(df)))]
    total = len(rows_to_process)
    skipped = 0
    errors = 0
    from_cache = 0

    for step, (idx, row) in enumerate(rows_to_process):
        name = get_cell_value(row, name_col)
        firma = get_cell_value(row, firma_col)
        titel = get_cell_value(row, titel_col)
        ort = get_cell_value(row, ort_col)

        if not name:
            skipped += 1
            progress_bar.progress((step + 1) / total)
            continue

        # Skip already enriched contacts (Excel-basiert)
        if skip_enriched and is_already_enriched(row):
            skipped += 1
            with status_container:
                st.markdown(f"⏭️ **{step+1}/{total}:** {name} — bereits angereichert, übersprungen")
            progress_bar.progress((step + 1) / total)
            continue

        # --- DB-Cache-Check: Bereits in Ergebnis-DB? ---
        if idx not in force_rerun:
            cached = enrichment_db.lookup(name, firma)
            if cached and cached.get("zusammenfassung"):
                from_cache += 1
                results[idx] = cached
                age = cached.get("_age_label", "")
                with status_container:
                    st.markdown(
                        f"💾 **{step+1}/{total}:** {name} — **aus DB geladen** ({age})"
                    )
                progress_bar.progress((step + 1) / total)
                continue

        with status_container:
            st.markdown(f"🔍 **{step+1}/{total}:** Recherchiere **{name}** ({firma})...")

        try:
            t_start = time.time()
            result = engine.enrich_contact(
                name=name,
                company=firma,
                title=titel,
                location=ort,
                max_searches=max_searches,
            )
            duration = round(time.time() - t_start, 1)
            results[idx] = result

            # --- AUTO-SAVE in DB (sofort, nicht erst beim Speichern-Button) ---
            try:
                enrichment_db.save(name, firma, result, metadata={
                    "title": titel,
                    "location": ort,
                    "llm_provider": engine.llm_provider,
                    "search_depth": max_searches,
                    "duration_seconds": duration,
                })
            except Exception as db_err:
                print(f"[DB SAVE] Warnung: {db_err}")

            email_info = result.get("email", "-") or "-"
            phone_info = result.get("personal_phone", "-") or "-"
            with status_container:
                st.markdown(
                    f"✅ **{name}** — "
                    f"E-Mail: `{email_info}` | "
                    f"Tel: `{phone_info}` | "
                    f"Status: {result.get('email_status', '-')} | "
                    f"⏱️ {duration}s · 💾 gespeichert"
                )

        except Exception as e:
            errors += 1
            # Generische Fehlermeldung im UI — keine API-Interna leaken
            err_type = type(e).__name__
            safe_msg = _sanitize_error(str(e))
            results[idx] = {"error": f"{err_type}: {safe_msg}"}
            with status_container:
                st.markdown(f"❌ **{name}** — Fehler beim Anreichern ({err_type})")
            # Voller Stacktrace nur in Server-Logs
            print(f"[ENRICH ERROR] {name}: {e}")

        progress_bar.progress((step + 1) / total)

    return results, skipped, errors, from_cache


def save_enriched_excel(df, results, output_path):
    """Speichert die angereicherte Excel-Datei mit Formatierung."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    enriched_df = df.copy()

    # Neue Spalten hinzufügen
    for col_name in NEW_COLUMNS:
        if col_name not in enriched_df.columns:
            enriched_df[col_name] = ""

    # Mapping von internen Keys zu Spaltennamen
    key_to_col = {
        "email": "E-Mail",
        "email_status": "E-Mail Status",
        "email_verifizierung": "E-Mail Verifizierung",
        "referenz_email": "Referenz-Email",
        "referenz_email_quelle": "Referenz-Email Quelle",
        "personal_phone": "Persönliches Telefon",
        "company_phone": "Firmentelefon (Impressum)",
        "conferences": "Konferenzen (2025/2026)",
        "podcasts_videos": "Podcasts / Videos",
        "job_changes": "Jobwechsel / Karriere",
        "linkedin_activity": "LinkedIn-Aktivität",
        "birthday": "Geburtstag",
        "relevant_info": "Relevante Infos für Ansprache",
        "zusammenfassung": "KI-Zusammenfassung",
        "personalisierte_nachricht": "Personalisierte Nachricht",
        "kanal_empfehlung": "Kanal-Empfehlung (Email/LinkedIn)",
        "monitoring_tags": "Monitoring-Tags (Delta-Scan)",
    }

    # Ergebnisse eintragen
    for idx, result in results.items():
        if "error" in result:
            enriched_df.at[idx, "Relevante Infos für Ansprache"] = f"FEHLER: {result['error']}"
            continue
        for key, col_name in key_to_col.items():
            val = result.get(key, "")
            if val:
                enriched_df.at[idx, col_name] = val
        enriched_df.at[idx, "Enrichment Datum"] = datetime.now().strftime("%Y-%m-%d")

    # Als Excel speichern
    enriched_df.to_excel(output_path, index=False, engine="openpyxl")

    # Formatierung anwenden
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # CI-Farben: Deep Purple #370C7B, Gold #FCB02F, Charcoal #2D2D2D
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill(start_color="370C7B", end_color="370C7B", fill_type="solid")
    new_col_fill = PatternFill(start_color="FCB02F", end_color="FCB02F", fill_type="solid")
    new_col_font = Font(bold=True, color="2D2D2D", size=11, name="Arial")
    enriched_fill = PatternFill(start_color="F3EEFB", end_color="F3EEFB", fill_type="solid")  # helles Purple
    error_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    original_col_count = len(df.columns)
    total_cols = len(enriched_df.columns)

    # Header formatieren
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        if col_idx > original_col_count:
            cell.fill = new_col_fill
            cell.font = new_col_font
        else:
            cell.fill = header_fill

    # Datenzeilen formatieren
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Neue Spalten grün markieren wenn befüllt
        row_data_idx = row_idx - 2  # 0-based index
        if row_data_idx in results:
            result = results[row_data_idx]
            fill = error_fill if "error" in result else enriched_fill
            for col_idx in range(original_col_count + 1, total_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell.fill = fill

    # Spaltenbreiten
    for col_idx in range(1, total_cols + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        if col_idx <= original_col_count:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 35

    # Name-Spalte breiter
    ws.column_dimensions["A"].width = 25

    # Freeze header + Auto-Filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(total_cols)}1"

    wb.save(output_path)
    return enriched_df


# ============================================================
# UI
# ============================================================

st.title("🧬 Contact Data Enrichment")
st.markdown('<div class="ci-accent"></div>', unsafe_allow_html=True)
st.markdown("**Selina Gaertner Consulting** — Lädt die Kontaktliste und reichert automatisch alle Kontakte an.")

# --- DSGVO-Hinweis (einmaliges Akzeptieren pro Session) ---
if "dsgvo_accepted" not in st.session_state:
    st.session_state.dsgvo_accepted = False

if not st.session_state.dsgvo_accepted:
    with st.expander("⚖️ Datenschutzhinweis (DSGVO) — bitte vor Nutzung lesen", expanded=True):
        st.markdown(f"""
**Verarbeitung personenbezogener Daten**

Diese App verarbeitet personenbezogene Daten (Namen, Firmen, Positionen, E-Mail-Adressen,
Telefonnummern) zum Zweck der **Geschäftsanbahnung (Lead-Generierung im B2B-Kontext)**.

**Was passiert mit den Daten:**
- **Lokale Verarbeitung:** Excel-Datei wird lokal auf deinem Rechner verarbeitet
- **API-Übertragung:** Name/Firma/Titel werden an Google Gemini und/oder Perplexity gesendet
  (Recherche-Zweck), sowie an Hunter.io (Email-Recherche)
- **Öffentliche Quellen:** Recherche erfolgt ausschließlich in öffentlich zugänglichen Quellen
  (Google, Unternehmenswebseiten, Presse, LinkedIn-öffentliche Profile)
- **Keine Speicherung bei Dritten:** Die APIs werden stateless genutzt — Daten werden nicht
  dauerhaft bei den LLM-Anbietern gespeichert (siehe deren Datenschutzrichtlinien)

**Deine Pflichten als Nutzerin/Nutzer (Art. 6 DSGVO):**
- Berechtigtes Interesse (Art. 6 Abs. 1 lit. f) muss gegeben sein
- Informationspflicht gegenüber dem Kontakt (Art. 14 DSGVO) bei erster Ansprache erfüllen
- Widerspruchsrecht respektieren, Daten bei Widerspruch löschen
- Angereicherte Daten **nur zweckgebunden** verwenden, nicht an Dritte weitergeben
- Aufbewahrung begrenzen — Leads nach abgeschlossenem Prozess löschen

**Verantwortlich:** Selina Gaertner Consulting · mail@selinagaertner.com
        """)
        dsgvo_col1, dsgvo_col2 = st.columns([1, 4])
        if dsgvo_col1.button("✓ Verstanden & akzeptiert", type="primary"):
            st.session_state.dsgvo_accepted = True
            st.rerun()
        dsgvo_col2.caption("Ohne Akzeptieren kann die App nicht genutzt werden.")
    st.stop()

# --- Sidebar ---
st.sidebar.title("⚙️ Einstellungen")

llm_choice = st.sidebar.radio(
    "🤖 KI-Modell",
    [
        "🚀 Gemini Deep Research (3-Stufen) — ~0,03 € / Kontakt",
        "🔬 Perplexity Deep Research — ~2–4 € / Kontakt",
    ],
    help=(
        "Gemini 3-Stufen: Plan → Search → Analyze. Flash entwirft gezielte Fragen, "
        "recherchiert mit Google Grounding, analysiert alles. ~25 Sek/Kontakt.\n\n"
        "Perplexity Deep Research: 30–90 Sek/Kontakt, 30+ Suchläufe, tiefste Dossiers."
    ),
)
llm_provider = "gemini" if "Gemini" in llm_choice else "perplexity"

if llm_provider == "gemini":
    api_key = st.sidebar.text_input(
        "Google Gemini API Key",
        value=os.environ.get("GEMINI_API_KEY", ""),
        type="password",
        help="Kostenlos via Google AI Studio (aistudio.google.com). Sehr günstiges Paid-Tier.",
    )
else:
    api_key = st.sidebar.text_input(
        "Perplexity API Key",
        value=os.environ.get("PERPLEXITY_API_KEY", ""),
        type="password",
        help="Via perplexity.ai → API → API Keys.",
    )

if not api_key:
    st.sidebar.warning("Ohne API Key: Keine KI-Zusammenfassung und keine personalisierten Nachrichten.")

hunter_api_key = st.sidebar.text_input(
    "Hunter.io API Key (optional)",
    value=os.environ.get("HUNTER_API_KEY", ""),
    type="password",
    help="Free Tier: 25 Suchen/Monat. Findet Email-Formate von Firmen. https://hunter.io",
)

search_depth = st.sidebar.selectbox(
    "Suchtiefe pro Kontakt",
    ["Schnell (3 Suchen)", "Normal (6 Suchen)", "Gründlich (10 Suchen)"],
    index=1,
)
depth_map = {"Schnell (3 Suchen)": 3, "Normal (6 Suchen)": 6, "Gründlich (10 Suchen)": 10}
max_searches = depth_map[search_depth]

skip_enriched = st.sidebar.checkbox(
    "Bereits angereicherte überspringen",
    value=True,
    help="Kontakte die schon eine E-Mail/Enrichment-Datum haben werden übersprungen.",
)

st.sidebar.markdown("---")

# Enrichment-DB Status + Export
_db_count = enrichment_db.count()
if _db_count > 0:
    st.sidebar.success(f"💾 **Ergebnis-DB:** {_db_count} Kontakte gespeichert")
    if st.sidebar.button("📥 Alle DB-Ergebnisse als Excel exportieren", key="export_db"):
        _export_path = os.path.join(
            DEFAULT_OUTPUT_DIR if os.path.isdir(DEFAULT_OUTPUT_DIR) else str(Path.home() / "Downloads"),
            f"enrichment_db_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        )
        try:
            enrichment_db.export_excel(_export_path)
            st.sidebar.success(f"✅ Exportiert: `{_export_path}`")
        except Exception as ex:
            st.sidebar.error(f"Export-Fehler: {_sanitize_error(str(ex))}")
else:
    st.sidebar.caption("💾 Ergebnis-DB: noch leer (wird beim ersten Enrichment befüllt)")

st.sidebar.markdown("---")
st.sidebar.markdown(
    f"**Geschätzte Zeit:** ~{max_searches * 5}s pro Kontakt\n\n"
    "DuckDuckGo hat Rate-Limits. Bei Fehlern: Suchtiefe reduzieren oder warten.\n\n"
    "**Hinweis:** Konstruierte E-Mails immer manuell prüfen!"
)

st.sidebar.markdown("---")

# LeadGen-DB-Status anzeigen
_leadgen_db = os.environ.get(
    "LEADGEN_DB_PATH",
    str(Path.home() / "Documents" / "AI Products I build" / "selina-ai-leadgen" / "data" / "automation.db"),
)
if os.path.exists(_leadgen_db):
    try:
        import sqlite3
        _conn = sqlite3.connect(f"file:{_leadgen_db}?mode=ro", uri=True, timeout=2.0)
        _count = _conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
        _conn.close()
        st.sidebar.success(f"🔗 **LeadGen-DB verbunden** · {_count} Kontakte werden abgeglichen")
    except Exception:
        st.sidebar.caption("🔗 LeadGen-DB gefunden, aber nicht lesbar")
else:
    st.sidebar.caption("⚪ Keine LeadGen-DB gefunden (kein Match-Check)")

st.sidebar.markdown("---")
st.sidebar.caption(
    "⚖️ **DSGVO-Hinweis:** Nur zweckgebundene B2B-Anbahnung. "
    "Informationspflicht (Art. 14) und Widerspruchsrecht respektieren. "
    "Angereicherte Daten nicht weitergeben."
)

# --- File Input ---
st.markdown("### 📂 Kontaktliste")

input_mode = st.radio(
    "Datei laden",
    ["Standard-Pfad verwenden", "Datei hochladen"],
    horizontal=True,
)

file_path = None
uploaded_df = None

def _validate_and_load_excel(source, source_name: str):
    """Lädt Excel mit Sicherheits-Checks: Size, Spaltenanzahl, Zeilenanzahl."""
    try:
        # Size-Check
        if hasattr(source, "size"):
            size_mb = source.size / (1024 * 1024)
        elif isinstance(source, str) and os.path.exists(source):
            size_mb = os.path.getsize(source) / (1024 * 1024)
        else:
            size_mb = 0

        if size_mb > MAX_UPLOAD_SIZE_MB:
            st.error(f"Datei zu groß ({size_mb:.1f} MB). Limit: {MAX_UPLOAD_SIZE_MB} MB.")
            return None

        # Excel laden
        df = pd.read_excel(source, engine="openpyxl")

        # Zeilen-Check
        if len(df) > MAX_CONTACTS:
            st.error(f"Zu viele Kontakte ({len(df)}). Limit: {MAX_CONTACTS}. Bitte Liste aufteilen.")
            return None

        # Plausibilitäts-Check: Muss mindestens eine Name-Spalte haben
        if find_column(df, NAME_COLS) is None:
            st.error("Keine 'Name'-Spalte gefunden. Erforderliche Spalten: Name, Firma, Titel.")
            return None

        return df
    except Exception:
        st.error(f"Datei konnte nicht geladen werden. Bitte prüfe das Excel-Format.")
        return None


if input_mode == "Standard-Pfad verwenden":
    file_path = st.text_input("Excel-Pfad", value=DEFAULT_EXCEL)
    if file_path and os.path.exists(file_path):
        uploaded_df = _validate_and_load_excel(file_path, file_path)
    elif file_path:
        st.warning(f"Datei nicht gefunden: {file_path}")
else:
    uploaded_file = st.file_uploader(
        f"Excel-Datei (.xlsx) — max {MAX_UPLOAD_SIZE_MB} MB, {MAX_CONTACTS} Kontakte",
        type=["xlsx"],
    )
    if uploaded_file:
        uploaded_df = _validate_and_load_excel(uploaded_file, uploaded_file.name)
        file_path = uploaded_file.name

# --- Kontakte anzeigen + starten ---
if uploaded_df is not None:
    df = uploaded_df
    st.session_state.df = df

    # Kontakt-Übersicht
    name_col = find_column(df, NAME_COLS)
    firma_col = find_column(df, FIRMA_COLS)
    titel_col = find_column(df, TITEL_COLS)

    already_enriched = sum(1 for _, row in df.iterrows() if is_already_enriched(row))

    # DB-Cache-Status für alle Kontakte prüfen
    _db_cache = {}
    for i in range(len(df)):
        _n = get_cell_value(df.iloc[i], name_col)
        _f = get_cell_value(df.iloc[i], firma_col) if firma_col else ""
        if _n:
            cached = enrichment_db.lookup(_n, _f)
            if cached and cached.get("zusammenfassung"):
                _db_cache[i] = cached

    col1, col2, col3 = st.columns(3)
    col1.metric("Kontakte gesamt", len(df))
    col2.metric("In Ergebnis-DB", len(_db_cache))
    col3.metric("Bereits in Excel", already_enriched)

    # --- Interaktive Auswahl-Tabelle ---
    st.markdown("### ✅ Kontakte auswählen")
    st.caption("Alle Kontakte sind vorausgewählt. Haken entfernen = wird nicht angereichert.")

    # Auswahl-Buttons
    btn_col1, btn_col2 = st.columns([1, 1])
    select_all = btn_col1.button("Alle auswählen", use_container_width=True)
    deselect_all = btn_col2.button("Alle abwählen", use_container_width=True)

    # Session State für Auswahl initialisieren
    if "contact_selection" not in st.session_state or len(st.session_state.contact_selection) != len(df):
        st.session_state.contact_selection = [True] * len(df)
    if select_all:
        st.session_state.contact_selection = [True] * len(df)
    if deselect_all:
        st.session_state.contact_selection = [False] * len(df)

    # Tabelle mit Checkboxen aufbauen
    show_cols = {}
    show_cols["Auswählen"] = st.session_state.contact_selection.copy()
    if name_col:
        show_cols["Name"] = [str(df.iloc[i][name_col]) if name_col in df.columns else "" for i in range(len(df))]
    if firma_col:
        show_cols["Firma"] = [str(df.iloc[i][firma_col]) if firma_col in df.columns else "" for i in range(len(df))]
    if titel_col:
        show_cols["Titel"] = [str(df.iloc[i][titel_col]) if titel_col in df.columns else "" for i in range(len(df))]
    # Status-Spalte falls vorhanden
    if "Status" in df.columns:
        show_cols["Status"] = [str(df.iloc[i]["Status"]) for i in range(len(df))]
    # Bereits angereichert markieren (Excel)
    show_cols["In Excel"] = ["✅" if is_already_enriched(df.iloc[i]) else "" for i in range(len(df))]
    # DB-Cache-Status anzeigen
    show_cols["In DB"] = [
        f"💾 {_db_cache[i]['_age_label']}" if i in _db_cache else ""
        for i in range(len(df))
    ]

    selection_df = pd.DataFrame(show_cols)

    edited = st.data_editor(
        selection_df,
        column_config={
            "Auswählen": st.column_config.CheckboxColumn("✓", default=True, width="small"),
            "Name": st.column_config.TextColumn("Name", width="medium"),
            "Firma": st.column_config.TextColumn("Firma", width="medium"),
            "Titel": st.column_config.TextColumn("Titel", width="large"),
            "Status": st.column_config.TextColumn("Status", width="small"),
            "In Excel": st.column_config.TextColumn("Excel", width="small"),
            "In DB": st.column_config.TextColumn("DB-Cache", width="medium"),
        },
        hide_index=True,
        use_container_width=True,
        disabled=[c for c in show_cols.keys() if c != "Auswählen"],
        key="contact_table",
    )

    # Ausgewählte Indizes ermitteln
    selected_indices = [i for i, checked in enumerate(edited["Auswählen"]) if checked]
    st.session_state.contact_selection = list(edited["Auswählen"])

    to_enrich = len(selected_indices)

    # --- START ---
    st.markdown("---")

    if to_enrich == 0:
        st.warning("Keine Kontakte ausgewählt. Bitte mindestens einen Kontakt auswählen.")
    else:
        estimated_time = to_enrich * max_searches * 2
        st.info(
            f"**{to_enrich} Kontakte** ausgewählt · **{max_searches} Suchen** pro Kontakt · "
            f"Geschätzte Dauer: ~{estimated_time // 60} Min {estimated_time % 60} Sek."
        )

    if st.button(f"🚀 {to_enrich} Kontakte anreichern", type="primary", disabled=(to_enrich == 0)):
        engine = EnrichmentEngine(
            api_key=api_key if api_key else None,
            hunter_api_key=hunter_api_key if hunter_api_key else None,
            llm_provider=llm_provider,
        )

        progress_bar = st.progress(0)
        status_container = st.container()

        # Welche Kontakte sollen trotz Cache neu recherchiert werden?
        force_rerun = {
            i for i in selected_indices
            if st.session_state.get(f"rerun_{i}", False)
        }

        with st.spinner("Enrichment läuft..."):
            results, skipped, errors, from_cache = run_enrichment(
                df, engine, max_searches, skip_enriched, progress_bar, status_container,
                selected_indices=selected_indices,
                force_rerun_indices=force_rerun,
            )

        # Force-Rerun-Flags zurücksetzen
        for i in force_rerun:
            st.session_state.pop(f"rerun_{i}", None)

        # --- Ergebnisse in Session State persistieren, damit Speichern-Button nach Rerun überlebt
        st.session_state.enrichment_results = results
        st.session_state.enrichment_skipped = skipped
        st.session_state.enrichment_errors = errors
        st.session_state.enrichment_from_cache = from_cache
        st.session_state.enrichment_timestamp = datetime.now().strftime('%Y%m%d_%H%M')

    # --- Zusammenfassung + Speichern: läuft AUSSERHALB des Anreichern-Buttons,
    # damit Speicher-Klicks einen Rerun überstehen
    if "enrichment_results" in st.session_state and st.session_state.enrichment_results:
        results = st.session_state.enrichment_results
        skipped = st.session_state.enrichment_skipped
        errors = st.session_state.enrichment_errors
        from_cache = st.session_state.get("enrichment_from_cache", 0)

        st.markdown("---")
        st.subheader("📊 Zusammenfassung")

        res_col1, res_col2, res_col3, res_col4, res_col5 = st.columns(5)
        res_col1.metric("Neu recherchiert", len(results) - errors - from_cache)
        res_col2.metric("Aus DB geladen", from_cache)
        res_col3.metric("Übersprungen", skipped)
        res_col4.metric("Fehler", errors)
        res_col5.metric("Gesamt", len(df))

        # --- Detail-Ergebnisse ---
        if results:
            # Übersicht: Anzahl LeadGen-Matches
            leadgen_matches = sum(
                1 for r in results.values()
                if isinstance(r, dict) and r.get("leadgen_match", {}).get("matched")
            )
            if leadgen_matches > 0:
                st.info(
                    f"🔗 **{leadgen_matches} von {len(results)}** Kontakten sind in deiner "
                    f"Lead-Gen-Historie — die Recherche berücksichtigt deren bisherigen Outbound-Verlauf."
                )

            with st.expander("🔎 Detail-Ergebnisse", expanded=True):
                for idx, result in results.items():
                    if "error" in result:
                        continue
                    row = df.iloc[idx]
                    name = get_cell_value(row, name_col)

                    # LeadGen-Match-Badge direkt neben dem Namen
                    lg = result.get("leadgen_match", {}) or {}
                    if lg.get("matched"):
                        status = lg.get("status", "")
                        confidence = lg.get("confidence", "")
                        st.markdown(f"#### {name} 🔗")
                        badge_color = CI_PURPLE
                        st.markdown(
                            f'<div style="background:{CI_LIGHT_GREY}; border-left:4px solid {badge_color}; '
                            f'padding:8px 12px; margin:4px 0 12px 0; border-radius:4px; font-family:Arial;">'
                            f'<strong style="color:{badge_color};">Bekannt aus Outbound</strong> · '
                            f'Status: <code>{status}</code> · Match-Konfidenz: {confidence}'
                            f'{"<br/>LinkedIn: <a href=\"" + lg["linkedin_url"] + "\" target=\"_blank\">" + lg["linkedin_url"] + "</a>" if lg.get("linkedin_url") else ""}'
                            f'{"<br/>Letzte Reply: <em>\"" + lg["reply_text"][:150] + "...\"</em>" if lg.get("reply_text") else ""}'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(f"#### {name}")
                    col_a, col_b = st.columns(2)
                    with col_a:
                        st.markdown(f"**E-Mail:** {result.get('email') or '-'}")
                        st.markdown(f"**Status:** {result.get('email_status') or '-'}")
                        verif = result.get('email_verifizierung') or '-'
                        st.markdown(f"**SMTP-Verifizierung:** {verif}")
                        ref = result.get('referenz_email') or '-'
                        ref_q = result.get('referenz_email_quelle') or ''
                        st.markdown(f"**Referenz-Email:** {ref}" + (f" ([Quelle]({ref_q}))" if ref_q else ""))
                        st.markdown(f"**Pers. Telefon:** {result.get('personal_phone') or '-'}")
                        st.markdown(f"**Firmentelefon:** {result.get('company_phone') or '-'}")
                    with col_b:
                        st.markdown(f"**Konferenzen:** {result.get('conferences') or '-'}")
                        st.markdown(f"**Podcasts/Videos:** {result.get('podcasts_videos') or '-'}")
                        st.markdown(f"**Jobwechsel:** {result.get('job_changes') or '-'}")
                        st.markdown(f"**LinkedIn:** {result.get('linkedin_activity') or '-'}")

                    # KI-Zusammenfassung + Nachricht prominent anzeigen
                    zusammenfassung = result.get('zusammenfassung') or ''
                    nachricht = result.get('personalisierte_nachricht') or ''
                    kanal = result.get('kanal_empfehlung') or ''

                    if zusammenfassung:
                        st.markdown("**📋 KI-Zusammenfassung:**")
                        st.info(zusammenfassung)

                    if nachricht:
                        kanal_icon = "📧" if "EMAIL" in kanal.upper() else "💬"
                        kanal_label = "Email" if "EMAIL" in kanal.upper() else "LinkedIn"
                        st.markdown(f"**{kanal_icon} Personalisierte Nachricht ({kanal_label}):**")
                        st.success(nachricht)

                    if kanal:
                        st.markdown(f"**🎯 Kanal-Empfehlung:** {kanal}")

                    st.markdown("---")

        # --- Export ---
        st.subheader("💾 Speichern")

        timestamp = st.session_state.get("enrichment_timestamp", datetime.now().strftime('%Y%m%d_%H%M'))
        output_name = f"kontakte_data_enriched_{timestamp}.xlsx"
        output_dir = DEFAULT_OUTPUT_DIR if os.path.isdir(DEFAULT_OUTPUT_DIR) else str(Path.home() / "Downloads")
        default_output_path = os.path.join(output_dir, output_name)

        output_path = st.text_input("Speicherpfad", value=default_output_path, key="save_path_input")

        if st.button("💾 Excel speichern", type="primary", key="save_button"):
            try:
                enriched_df = save_enriched_excel(df, results, output_path)
                st.success(f"✅ Gespeichert unter: `{output_path}`")
                st.balloons()
            except PermissionError:
                st.error(
                    f"❌ **Keine Schreibrechte** für diesen Pfad:\n`{output_path}`\n\n"
                    "→ Bitte anderen Ordner wählen (z.B. `~/Documents/` oder `~/Downloads/`) "
                    "oder prüfen, ob die Datei in Excel geöffnet ist."
                )
            except FileNotFoundError:
                st.error(
                    f"❌ **Ordner existiert nicht:** `{os.path.dirname(output_path)}`\n\n"
                    "→ Bitte gültigen Pfad angeben."
                )
            except Exception as e:
                err_type = type(e).__name__
                safe_msg = _sanitize_error(str(e))
                st.error(f"❌ **Fehler beim Speichern ({err_type}):** {safe_msg}")
                print(f"[SAVE ERROR] {output_path}: {e}")

        # Download-Button
        buffer = BytesIO()
        temp_df = df.copy()
        for col_name in NEW_COLUMNS:
            if col_name not in temp_df.columns:
                temp_df[col_name] = ""
        key_to_col = {
            "email": "E-Mail", "email_status": "E-Mail Status",
            "email_verifizierung": "E-Mail Verifizierung",
            "referenz_email": "Referenz-Email", "referenz_email_quelle": "Referenz-Email Quelle",
            "personal_phone": "Persönliches Telefon", "company_phone": "Firmentelefon (Impressum)",
            "conferences": "Konferenzen (2025/2026)", "podcasts_videos": "Podcasts / Videos",
            "job_changes": "Jobwechsel / Karriere", "linkedin_activity": "LinkedIn-Aktivität",
            "birthday": "Geburtstag", "relevant_info": "Relevante Infos für Ansprache",
            "zusammenfassung": "KI-Zusammenfassung",
            "personalisierte_nachricht": "Personalisierte Nachricht",
            "kanal_empfehlung": "Kanal-Empfehlung (Email/LinkedIn)",
        }
        for idx, result in results.items():
            if "error" in result:
                continue
            for key, col_name in key_to_col.items():
                val = result.get(key, "")
                if val:
                    temp_df.at[idx, col_name] = val
            temp_df.at[idx, "Enrichment Datum"] = datetime.now().strftime("%Y-%m-%d")
        temp_df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)

        st.download_button(
            label="📥 Download Excel",
            data=buffer,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

else:
    st.markdown("""
    ### So funktioniert's:
    1. **Excel laden** — Standard-Pfad oder Datei hochladen
    2. **Start klicken** — App geht automatisch ALLE Kontakte durch
    3. **Für jeden Kontakt wird gesucht nach:**
       - 📧 E-Mail-Adresse (verifiziert oder aus Firmenformat konstruiert)
       - 📞 Persönliches Telefon + Firmentelefon aus Impressum
       - 🎤 Konferenzen, Podcasts, YouTube-Videos
       - 💼 Jobwechsel, LinkedIn-Aktivität
       - 🎂 Geburtstag und relevante Infos für die Ansprache
    4. **Ergebnis** — Neue Excel-Datei mit allen Enrichment-Spalten
    """)

# --- CI Footer ---
st.markdown(
    '<div class="ci-footer">'
    'Selina Gaertner Consulting · AI Strategy & Governance for Life Sciences & MedTech<br>'
    'selinagaertner.com · mail@selinagaertner.com'
    '</div>',
    unsafe_allow_html=True,
)
